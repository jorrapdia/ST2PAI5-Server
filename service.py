from Crypto.PublicKey import RSA
from Crypto.Signature import PKCS1_v1_5
from Crypto.Hash import SHA512
from datetime import datetime, timedelta
import os
import openpyxl
import xlsxwriter
import db as database
import dateutil.relativedelta

from openpyxl.styles import Side, Border


def validate_order(quantities):
    res = True
    for quantity in quantities:
        if int(quantity) < 0 or int(quantity) > 300:
            res = False
    return res


def verify_user(username, public_key):
    res = True
    public_key_db = database.get_public_key_by_client_number(username)
    if public_key_db is None:
        database.insert_user(username, public_key)
    else:
        res = public_key_db == public_key
    return res


def verify_signature(order, signature, public_key_pem):
    public_key = RSA.import_key(bytearray.fromhex(public_key_pem))
    signer = PKCS1_v1_5.new(public_key)
    digest = SHA512.new()
    digest.update(order.encode())
    return signer.verify(digest, bytearray.fromhex(signature))


def kpi():
    db = database.get_db()
    cursor = db.cursor()
    now = datetime.now()
    p3_query = "SELECT COUNT(*) FROM orders WHERE (order_date >= " + str(
        (now - dateutil.relativedelta.relativedelta(months=2)).timestamp()) + " AND created_at <= " + str(
        now.timestamp()) + ")"
    p2_query = "SELECT COUNT(*) FROM orders WHERE (order_date >= " + str(
        (now - dateutil.relativedelta.relativedelta(months=3)).timestamp()) + " AND created_at <= " + str(
        (now - dateutil.relativedelta.relativedelta(months=1)).timestamp()) + ")"
    p1_query = "SELECT COUNT(*) FROM orders WHERE (order_date >= " + str(
        (now - dateutil.relativedelta.relativedelta(months=4)).timestamp()) + " AND created_at <= " + str(
        (now - dateutil.relativedelta.relativedelta(months=2)).timestamp()) + ")"

    total_query = "SELECT COUNT(*) FROM orders"
    p3 = cursor.execute(p3_query).fetchone()[0]
    p2 = cursor.execute(p2_query).fetchone()[0]
    p1 = cursor.execute(p1_query).fetchone()[0]

    total = cursor.execute(total_query).fetchone()[0]

    p3_ratio = (p3 / total) * 100
    p2_ratio = (p2 / total) * 100
    p1_ratio = (p1 / total) * 100

    if (p3_ratio > p1_ratio and p3_ratio > p2_ratio) or (p3_ratio > p1_ratio and p3_ratio == p2_ratio) or (
            p3_ratio == p1_ratio and p3_ratio > p2_ratio):
        tendencia = '+'
    elif p3_ratio < p1_ratio or p3_ratio < p2_ratio:
        tendencia = '-'
    else:
        tendencia = '0'

    return p3_ratio, tendencia


def update_kpi():
    dirname = os.path.dirname(__file__)
    filename = 'KPI.xlsx'
    pathname = os.path.join(dirname, filename)
    ratio, tendencia = kpi()

    if os.path.exists(pathname):
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook['KPI']
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        max_row = worksheet.max_row

        worksheet.cell(row=max_row + 1, column=2).value = datetime.now().strftime("%B (%Y)")
        worksheet.cell(row=max_row + 1, column=2).border = thin_border
        worksheet.cell(row=max_row + 1, column=3).value = str(ratio) + '%'
        worksheet.cell(row=max_row + 1, column=3).border = thin_border
        worksheet.cell(row=max_row + 1, column=4).value = tendencia
        worksheet.cell(row=max_row + 1, column=4).border = thin_border
        workbook.save(filename)

    else:
        workbook = xlsxwriter.Workbook(pathname)
        worksheet = workbook.add_worksheet('KPI')

        title_format = workbook.add_format(
            {'bold': 1, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': '#000000'})
        column_title_format = workbook.add_format(
            {'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': '#000000'})
        data_format = workbook.add_format({'valign': 'vcenter', 'border': 1, 'border_color': '#000000'})

        worksheet.set_column(1, 3, 24)

        worksheet.merge_range("B2:G2", 'KPI Hotel', title_format)
        worksheet.write(2, 1, 'Fecha', column_title_format)
        worksheet.write(3, 1, datetime.now().strftime("%d/%m/%Y %H:%M:%S"), data_format)
        worksheet.write(2, 2, 'Ratio', column_title_format)
        worksheet.write(3, 2, str(ratio) + '%', data_format)
        worksheet.write(2, 3, 'Tendencia', column_title_format)
        worksheet.write(3, 3, tendencia, data_format)
        workbook.close()
    print('SERVER INFO: KPI was updated')