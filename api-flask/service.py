from db import get_db
from Crypto.PublicKey import RSA 
from Crypto.Signature import PKCS1_v1_5 
from Crypto.Hash import SHA256 
from base64 import b64decode
from datetime import datetime, timedelta
import os
import openpyxl
import xlsxwriter

from openpyxl.styles import Side, Border


def insert_order(beds, tables, chairs, armchairs, client_number, created_at, verify):

    db = get_db()
    cursor = db.cursor()
    statement = 'INSERT INTO orders(beds, tables, chairs, armchairs, client_number, created_at, verify) VALUES (?, ?, ?, ?, ?, ?, ?)'
    cursor.execute(statement, [beds, tables, chairs, armchairs, client_number, created_at, verify])
    db.commit()

def get_public_key_by_client_number(client_number):

    db = get_db()
    cursor = db.cursor()
    statement = "SELECT public_key FROM employeers WHERE client_number = ?"
    cursor.execute(statement, [client_number])

    return cursor.fetchone()

def validate_order(order):

    res = True
    for key in order.keys():
        if order[key] < 0 | order[key] > 300:
            res = False
    
    return res

def verify_signature(order, signature, client_number):

    public_key_pem = get_public_key_by_client_number(client_number)
    public_key = RSA.PublicKey.load_pkcs1(public_key_pem.encode('utf8')) 
    signer = PKCS1_v1_5.new(public_key)
    digest = SHA256.new()
    digest.update(b64decode(order))

    return signer.verify(digest, b64decode(signature))

def KPI():

    db = get_db()
    cursor = db.cursor()

    now =  datetime.now()

    p3_query = "SELECT COUNT(*) FROM orders WHERE (created_at >=" + (now - timedelta(months=2)).strftime('%d/%m/%Y, %H:%M:%S') + "AND created_at <=" + now.strftime('%d/%m/%Y, %H:%M:%S') + ") AND verify = 'True'"
    p2_query = "SELECT COUNT(*) FROM orders WHERE (created_at >=" + (now - timedelta(months=3)).strftime('%d/%m/%Y, %H:%M:%S') + "AND created_at <=" + (now - timedelta(months=1)).strftime('%d/%m/%Y, %H:%M:%S') + ") AND verify = 'True'"
    p1_query = "SELECT COUNT(*) FROM orders WHERE (created_at >=" + (now - timedelta(months=4)).strftime('%d/%m/%Y, %H:%M:%S') + "AND created_at <=" + (now - timedelta(months=2)).strftime('%d/%m/%Y, %H:%M:%S') + ") AND verify = 'True'"
    
    total_query = "SELECT COUNT(*) FROM orders"

    p3 = cursor.execute(p3_query).fetchone()[0]
    p2 = cursor.execute(p2_query).fetchone()[0]
    p1 = cursor.execute(p1_query).fetchone()[0]

    total = cursor.execute(total_query).fetchone()[0]

    p3_ratio = (p3/total) * 100
    p2_ratio = (p2/total) * 100
    p1_ratio = (p1/total) * 100


    if(p3_ratio > p1_ratio and p3_ratio > p2_ratio) or (p3_ratio > p1_ratio and p3_ratio == p2_ratio) or (p3_ratio == p1_ratio and p3_ratio > p2_ratio):
        tendencia = '+'
    elif (p3_ratio < p1_ratio or p3_ratio < p2_ratio):
        tendencia = '-'
    else:
        tendencia = '0'

    return p3_ratio, tendencia

def update_KPI():
    dirname = os.path.dirname(__file__)
    filename = 'KPI.xlsx'
    pathname = os.path.join(dirname, filename)

    ratio, tendencia = KPI()

    if os.path.exists(pathname):
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook['KPI']
        thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
        max_row = worksheet.max_row

        worksheet.cell(row=max_row+1, column=2).value = datetime.now().strftime("%B (%Y)")
        worksheet.cell(row=max_row+1, column=2).border = thin_border
        worksheet.cell(row=max_row+1, column=3).value = str(ratio) + '%'
        worksheet.cell(row=max_row+1, column=3).border = thin_border
        worksheet.cell(row=max_row+1, column=4).value = tendencia
        worksheet.cell(row=max_row+1, column=4).border = thin_border

        workbook.save(filename)

    else:
        workbook = xlsxwriter.Workbook(pathname)
        worksheet = workbook.add_worksheet('KPI')

        title_format = workbook.add_format({'bold': 1, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': '#000000'})
        column_title_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': '#000000'})
        data_format = workbook.add_format({'valign': 'vcenter', 'border': 1, 'border_color': '#000000'})

        worksheet.set_column(1, 3, 24)

        worksheet.merge_range("B2:G2", 'KPI Hotel', title_format)
        worksheet.write(2, 1, 'Fecha', column_title_format)
        worksheet.write(3, 1, datetime.now().strftime("%d/%m/%Y %H:%M:%S"), data_format)

        worksheet.write(2, 2, 'Ratio', column_title_format)
        worksheet.write(3, 2, str(ratio) + '%', data_format)

        worksheet.write(2, 3, 'Tendencia', column_title_format)
        worksheet.write(3, 3, tendencia, data_format)

        workbook.close()
    print('SERVER INFO: KPI was updated')